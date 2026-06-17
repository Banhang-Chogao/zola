"""Parse VietinBank Internet Banking account statement Excel exports."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SOURCE = "vietinbank"

HEADER_MARKERS = {
    "stt": ("stt",),
    "date": ("ngày", "ngay"),
    "description": ("nội dung", "noi dung"),
    "amount": ("số tiền gd", "so tien gd"),
    "balance": ("số dư", "so du"),
}

DATE_FORMATS = (
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
)


def _norm(text: Any) -> str:
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_number(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(value))

    s = str(value).strip()
    s = s.replace(",", "").replace(" ", "")
    if not s:
        return None

    sign = 1
    if s.startswith("+"):
        s = s[1:]
    elif s.startswith("-"):
        sign = -1
        s = s[1:]

    if not s.replace(".", "", 1).isdigit():
        return None
    return sign * int(round(float(s)))


def parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%S")

    s = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%S")
        except ValueError:
            continue
    return None


def font_color_hint(cell) -> str | None:
    """Return 'expense' (red) or 'income' (green/blue) when font color is present."""
    if cell is None or cell.font is None or cell.font.color is None:
        return None

    color = cell.font.color
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return None

    rgb = rgb.upper()
    if len(rgb) == 8:
        rgb = rgb[2:]

    if len(rgb) != 6:
        return None

    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    if r > 180 and g < 100 and b < 100:
        return "expense"
    if g > 100 and r < 150:
        return "income"
    if b > 150 and r < 100:
        return "income"
    return None


def amount_type(amount: int, color_hint: str | None) -> str:
    if amount < 0:
        return "expense"
    if amount > 0:
        return "income"
    if color_hint in ("expense", "income"):
        return color_hint
    return "expense"


def transaction_id(date: str, description: str, amount: int, balance: int) -> str:
    payload = f"{date}|{description}|{amount}|{balance}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _row_cells(ws: Worksheet, row_idx: int) -> list[Any]:
    max_col = ws.max_column or 10
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


def _find_header_row(ws: Worksheet) -> tuple[int, dict[str, int]] | None:
    for row_idx in range(1, (ws.max_row or 0) + 1):
        cells = _row_cells(ws, row_idx)
        norms = [_norm(c) for c in cells]
        cols: dict[str, int] = {}

        for col_idx, text in enumerate(norms):
            if not text:
                continue
            for key, markers in HEADER_MARKERS.items():
                if key in cols:
                    continue
                if any(m in text for m in markers):
                    cols[key] = col_idx + 1

        if all(k in cols for k in HEADER_MARKERS):
            return row_idx, cols
    return None


def _cell(ws: Worksheet, row: int, col: int):
    return ws.cell(row=row, column=col)


def parse_vietinbank(
    path: str | Path,
    *,
    sheet_index: int = 0,
) -> list[dict[str, Any]]:
    """Parse a VietinBank statement .xlsx into normalized transaction dicts."""
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.worksheets[sheet_index]

    found = _find_header_row(ws)
    if not found:
        wb.close()
        raise ValueError(
            "Không tìm thấy dòng tiêu đề bảng (STT / Ngày / Nội dung / Số tiền GD / Số dư)"
        )

    header_row, cols = found
    transactions: list[dict[str, Any]] = []

    for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
        stt = _cell(ws, row_idx, cols["stt"]).value
        if stt is None or str(stt).strip() == "":
            continue

        date_raw = _cell(ws, row_idx, cols["date"]).value
        description = _cell(ws, row_idx, cols["description"]).value
        amount_cell = _cell(ws, row_idx, cols["amount"])
        balance_raw = _cell(ws, row_idx, cols["balance"]).value

        date_iso = parse_date(date_raw)
        amount = parse_number(amount_cell.value)
        balance = parse_number(balance_raw)

        if date_iso is None or amount is None or balance is None:
            continue

        desc = str(description).strip() if description is not None else ""
        color_hint = font_color_hint(amount_cell)
        tx_type = amount_type(amount, color_hint)
        tx_id = transaction_id(date_iso, desc, amount, balance)

        transactions.append(
            {
                "transaction_id": tx_id,
                "date": date_iso,
                "description": desc,
                "amount": amount,
                "type": tx_type,
                "balance": balance,
                "source": SOURCE,
            }
        )

    wb.close()
    return transactions


def merge_transactions(
    parsed: list[dict[str, Any]],
    existing_ids: set[str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Split parsed rows into INSERT (new) and SKIP (duplicate transaction_id)."""
    seen = set(existing_ids or ())
    inserted: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for tx in parsed:
        tx_id = tx["transaction_id"]
        if tx_id in seen:
            skipped.append(tx)
        else:
            inserted.append(tx)
            seen.add(tx_id)

    return inserted, skipped


def main() -> int:
    parser = argparse.ArgumentParser(description="Parse VietinBank Excel statement to JSON")
    parser.add_argument("xlsx", type=Path, help="Path to .xlsx statement file")
    parser.add_argument(
        "--existing-ids",
        type=Path,
        help="JSON file with list of existing transaction_id strings",
    )
    args = parser.parse_args()

    rows = parse_vietinbank(args.xlsx)
    existing: set[str] = set()
    if args.existing_ids and args.existing_ids.exists():
        data = json.loads(args.existing_ids.read_text(encoding="utf-8"))
        existing = set(data if isinstance(data, list) else data.get("ids", []))

    inserted, skipped = merge_transactions(rows, existing)
    out = {
        "total_parsed": len(rows),
        "inserted": len(inserted),
        "skipped": len(skipped),
        "transactions": rows,
        "new_transactions": inserted,
    }
    json.dump(out, sys.stdout, ensure_ascii=False, indent=2)
    print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())