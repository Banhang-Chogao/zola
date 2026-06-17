"""Tests for F-Dashboard VietinBank parser and insights."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from f_dashboard_insights import build_insights_payload, categorize_expense
from f_dashboard_parse_excel import merge_transactions as merge_tx
from f_dashboard_parse_excel import parse_vietinbank, transaction_id as tx_id_fn

SAMPLE_ROWS = [
    (1, "17/06/2026 20:22:44", "QR - NGUYEN DUY KHANG chuyen tien", "-200,000", "6,441,308", "FF0000"),
    (2, "17/06/2026 19:00:56", "QR - NGUYEN DUY KHANG chuyen tien", "-750,000", "6,641,308", "FF0000"),
    (3, "13/06/2026 16:44:15", "NGUYEN DUY KHANG chuyen tien", "-7,391,308", "7,391,308", "FF0000"),
    (28, "09/06/2026 10:59:37", "MBVCB.14588949863...", "+78,000", "178,000", "008000"),
    (31, "09/06/2026 10:19:54", "MBVCB.14588425322...", "+65,000", "65,000", "008000"),
]


def _build_sample_xlsx(path: Path, *, with_colors: bool = True) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sao ke"

    ws["A1"] = "VietinBank"
    ws["A2"] = "Sao kê tài khoản, Số tài khoản: 106887370108 - NGUYEN DUY KHANG - Tài khoản thanh toán"
    ws["A3"] = "Từ ngày 18/05/2026 đến ngày 17/06/2026"
    ws["A4"] = "Loại tiền: VND"

    headers = ["STT", "Ngày", "Nội dung", "Số tiền GD", "Số dư"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=6, column=col, value=h)

    for i, (stt, dt, desc, amt, bal, color) in enumerate(SAMPLE_ROWS, start=7):
        ws.cell(row=i, column=1, value=stt)
        ws.cell(row=i, column=2, value=dt)
        ws.cell(row=i, column=3, value=desc)
        amt_cell = ws.cell(row=i, column=4, value=amt)
        ws.cell(row=i, column=5, value=bal)
        if with_colors:
            amt_cell.font = Font(color=color)

    wb.save(path)
    wb.close()


class VietinBankParserTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.sample = Path(self.tmp.name) / "vietinbank_sample.xlsx"
        _build_sample_xlsx(self.sample)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_skips_header_and_parses_five_transactions(self) -> None:
        rows = parse_vietinbank(self.sample)
        self.assertEqual(len(rows), 5)

    def test_expense_and_income_by_sign(self) -> None:
        plain = Path(self.tmp.name) / "plain.xlsx"
        _build_sample_xlsx(plain, with_colors=False)
        rows = parse_vietinbank(plain)
        expenses = [r for r in rows if r["type"] == "expense"]
        incomes = [r for r in rows if r["type"] == "income"]
        self.assertEqual(len(expenses), 3)
        self.assertEqual(len(incomes), 2)

    def test_transaction_id_stable(self) -> None:
        expected = tx_id_fn(
            "2026-06-17T20:22:44",
            "QR - NGUYEN DUY KHANG chuyen tien",
            -200000,
            6441308,
        )
        rows = parse_vietinbank(self.sample)
        self.assertEqual(rows[0]["transaction_id"], expected)
        self.assertEqual(len(rows[0]["transaction_id"]), 64)

    def test_dedup_skip_insert(self) -> None:
        rows = parse_vietinbank(self.sample)
        existing = {rows[0]["transaction_id"], rows[1]["transaction_id"]}
        inserted, skipped = merge_tx(rows, existing)
        self.assertEqual(len(inserted), 3)
        self.assertEqual(len(skipped), 2)


class InsightsTest(unittest.TestCase):
    def test_build_insights_payload(self) -> None:
        txs = [
            {"transaction_id": "a", "date": "2026-06-17T20:22:44", "description": "Starbucks coffee", "amount": -50000, "type": "expense", "balance": 100000, "source": "vietinbank"},
            {"transaction_id": "b", "date": "2026-06-17T10:00:00", "description": "Salary", "amount": 500000, "type": "income", "balance": 150000, "source": "vietinbank"},
        ]
        payload = build_insights_payload(txs)
        self.assertIn("summary", payload)
        self.assertIn("health", payload)
        self.assertIn("charts", payload)
        self.assertIn("insights", payload)
        self.assertEqual(payload["summary"]["total_income"], 500000)
        self.assertEqual(payload["summary"]["total_expense"], 50000)
        self.assertGreater(payload["health"]["financial_score"], 0)

    def test_categorize_expense(self) -> None:
        self.assertEqual(categorize_expense("STARBUCKS Nguyen Hue"), "Ăn uống")
        self.assertEqual(categorize_expense("QR - chuyen tien"), "Chuyển tiền")


class CliTest(unittest.TestCase):
    def test_parse_cli(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            sample = Path(tmp) / "sample.xlsx"
            _build_sample_xlsx(sample)
            script = Path(__file__).resolve().parent / "f_dashboard_parse_excel.py"
            result = subprocess.run(
                [sys.executable, str(script), str(sample)],
                capture_output=True,
                text=True,
                check=True,
            )
            data = json.loads(result.stdout)
            self.assertEqual(data["total_parsed"], 5)


if __name__ == "__main__":
    unittest.main()